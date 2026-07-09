import json
import sys
import openpyxl
from datetime import datetime

SRC = "/root/.claude/uploads/0cacce55-a6bc-5b51-b763-2e52695ef53e/7c30cb63-C_pia_de_Controle_de_medi__o_Mensal__Atualizado_at__Junho_26.xlsx"
OUT = "/workspace/controle-medicoes-contratos-obra-epr-br153-lote2/src/data/seed/obra-156-epr-br153-lote2.json"

MONTH_COLS = {
    "Q": "2025-12", "R": "2026-01", "S": "2026-02", "T": "2026-03",
    "U": "2026-04", "V": "2026-05", "W": "2026-06", "X": "2026-07",
    "Y": "2026-08", "Z": "2026-09", "AA": "2026-10", "AB": "2026-11",
    "AC": "2026-12",
}

FILIAIS = {
    156: {"codigo": 156, "sigla": "PV", "nome": "156 - PV"},
    157: {"codigo": 157, "sigla": "FD", "nome": "157 - FD (Faturamento Direto)"},
}

def num(v):
    if v is None:
        return 0
    return round(float(v), 2)

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["CONTRATO - MEDIÇÕES"]

    contratos = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        cells = {c.coordinate[0] if not c.coordinate[1].isalpha() else c.column_letter: c.value for c in row}
        by_col = {c.column_letter: c.value for c in row}
        contrato_id = by_col.get("B")
        if not contrato_id:
            continue

        filial = by_col.get("D")
        data_inicio = by_col.get("H")
        if isinstance(data_inicio, datetime):
            data_inicio = data_inicio.date().isoformat()

        medicoes = []
        for col, mes in MONTH_COLS.items():
            v = by_col.get(col)
            if v is not None and v != 0:
                medicoes.append({"mes": mes, "valor": num(v)})

        contratos.append({
            "id": str(contrato_id).strip(),
            "tipoContrato": (by_col.get("C") or "").strip(),
            "filial": filial,
            "codigoFornecedor": by_col.get("E"),
            "fornecedor": (by_col.get("F") or "").strip(),
            "objetivo": (by_col.get("G") or "").strip(),
            "dataInicio": data_inicio,
            "valorContrato": num(by_col.get("I")),
            "valorAditivo": num(by_col.get("J")),
            "valorTotal": num(by_col.get("K")),
            "saldo": num(by_col.get("M")),
            "valorMedido": num(by_col.get("O")),
            "medicoes": medicoes,
            "planejado": [],
        })

    obra = {
        "id": "obra-156-epr-br153-lote2",
        "nome": "OBRA 156 - EPR DUPLICAÇÃO BR 153 LOTE 2",
        "descricao": "Controle de Contratos / Medições",
        "filiais": list(FILIAIS.values()),
        "atualizadoEm": datetime.now().isoformat(),
        "contratos": contratos,
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(obra, f, ensure_ascii=False, indent=2)

    print(f"OK: {len(contratos)} contratos exportados para {OUT}")

if __name__ == "__main__":
    main()
